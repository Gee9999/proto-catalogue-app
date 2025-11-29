import io
import re
import tempfile
from typing import Optional, Dict, Any

import pandas as pd
from PIL import Image

import streamlit as st
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage


# -----------------------------
# Helpers: column detection
# -----------------------------

def find_code_col(df: pd.DataFrame) -> str:
    """Try to find the product code column in a generic Excel file."""
    cols = list(df.columns)

    # 1) Exact match
    for col in cols:
        if str(col).strip().lower() == "code":
            return col

    # 2) Contains word "code" or "barcode" or "item"
    for col in cols:
        name = str(col).strip().lower()
        if "barcode" in name or "code" in name or "item" in name:
            return col

    raise ValueError("Could not find a CODE column (e.g. 'CODE', 'Barcode', 'ItemCode').")


def find_desc_col(df: pd.DataFrame) -> str:
    """Try to find the description column."""
    cols = list(df.columns)

    # 1) Exact "description"
    for col in cols:
        if str(col).strip().lower() == "description":
            return col

    # 2) Contains desc/name/product
    for col in cols:
        name = str(col).strip().lower()
        if "description" in name or "desc" in name or "name" in name or "product" in name:
            return col

    raise ValueError("Could not find a DESCRIPTION column (e.g. 'DESCRIPTION', 'Product Name').")


def find_price_incl_col(df: pd.DataFrame) -> str:
    """Try to find the VAT-inclusive price column."""
    cols = list(df.columns)

    # 1) Very likely names: includes 'incl' & ('price' or 'vat')
    for col in cols:
        name = str(col).strip().lower()
        if "incl" in name and ("price" in name or "vat" in name):
            return col

    # 2) Any column with 'price a incl' / 'price-a incl' etc.
    for col in cols:
        name = str(col).strip().lower().replace(" ", "").replace("-", "")
        if "priceaincl" in name or "priceincl" in name:
            return col

    # 3) Fallback: first column containing 'price'
    for col in cols:
        name = str(col).strip().lower()
        if "price" in name:
            return col

    raise ValueError("Could not find a VAT-inclusive price column (e.g. 'PRICE-A INCL').")


# -----------------------------
# Helpers: code extraction & normalisation
# -----------------------------

def extract_code_from_filename(filename: str) -> Optional[str]:
    """
    Extract the first long (6+ digit) number from the photo filename.
    This is your product code by Option 1 rule.
    """
    base = filename.rsplit(".", 1)[0]
    match = re.search(r"\d{6,}", base)
    if not match:
        return None
    return match.group(0)


def normalize_code_from_excel(val: Any) -> Optional[str]:
    """
    Normalise the CODE from Excel to the same style as filenames:
    - Try first long digit sequence (6+ digits)
    - Fallback: strip non-digits
    """
    if pd.isna(val):
        return None
    s = str(val).strip()
    # Prefer a long block of digits, same as filenames
    m = re.search(r"\d{6,}", s)
    if m:
        return m.group(0)
    # Fallback: all digits
    digits = re.sub(r"[^\d]", "", s)
    return digits if digits else None


# -----------------------------
# Build matched DataFrame
# -----------------------------

def build_matched_df(price_df: pd.DataFrame,
                     uploaded_photos: list,
                     code_col: str,
                     desc_col: str,
                     price_col: str) -> pd.DataFrame:
    """
    For each uploaded photo, extract code from filename,
    match exactly to Excel, and return a DataFrame with:
      FILENAME, FILE, CODE, DESCRIPTION, PRICE_INCL
    """
    # Create a normalized key column for codes in the Excel
    price_df = price_df.copy()
    price_df["CODE_KEY"] = price_df[code_col].apply(normalize_code_from_excel)

    # Build a lookup dict: code_key -> row dict
    code_to_row: Dict[str, Dict[str, Any]] = {}
    for _, row in price_df.iterrows():
        key = row["CODE_KEY"]
        if key:
            # If duplicates, last one wins – fine for your catalogue
            code_to_row[key] = {
                "CODE": row[code_col],
                "DESCRIPTION": row[desc_col],
                "PRICE_INCL": row[price_col],
            }

    records = []

    for file in uploaded_photos:
        filename = file.name
        extracted_code = extract_code_from_filename(filename)
        rec_code = ""
        rec_desc = ""
        rec_price = ""

        if extracted_code and extracted_code in code_to_row:
            data = code_to_row[extracted_code]
            rec_code = str(data["CODE"])
            rec_desc = str(data["DESCRIPTION"])
            rec_price = data["PRICE_INCL"]
        else:
            # No match found
            rec_code = extracted_code or ""
            rec_desc = ""
            rec_price = ""

        records.append({
            "FILENAME": filename,
            "FILE": file,
            "CODE": rec_code,
            "DESCRIPTION": rec_desc,
            "PRICE_INCL": rec_price,
        })

    return pd.DataFrame(records)


# -----------------------------
# Excel with thumbnails
# -----------------------------

def build_excel_with_thumbnails(df: pd.DataFrame, temp_dir: str) -> bytes:
    """
    Create an Excel file with:
      Col A: Photo thumbnail
      Col B: CODE
      Col C: DESCRIPTION
      Col D: PRICE_INCL
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalogue"

    headers = ["PHOTO", "CODE", "DESCRIPTION", "PRICE_INCL"]
    ws.append(headers)

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 15

    row_idx = 2

    for _, row in df.iterrows():
        # Set a decent row height for thumbnails
        ws.row_dimensions[row_idx].height = 80

        # Write text cells
        ws.cell(row=row_idx, column=2, value=row["CODE"])
        ws.cell(row=row_idx, column=3, value=row["DESCRIPTION"])
        ws.cell(row=row_idx, column=4, value=row["PRICE_INCL"])

        # Add thumbnail in column A
        file = row["FILE"]
        try:
            img = Image.open(file)
            img.thumbnail((120, 120))

            tmp = tempfile.NamedTemporaryFile(
                dir=temp_dir, suffix=".png", delete=False
            )
            tmp_path = tmp.name
            tmp.close()  # close handle so Excel can re-open it
            img.save(tmp_path, format="PNG")

            xl_img = XLImage(tmp_path)
            xl_img.anchor = f"A{row_idx}"
            ws.add_image(xl_img)
        except Exception:
            # If image fails, leave cell blank
            pass

        row_idx += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# -----------------------------
# PDF 3×3 grid builder
# -----------------------------

def build_pdf_grid(df: pd.DataFrame, temp_dir: str) -> bytes:
    """
    Build a 3x3 grid PDF:
      - Photo
      - Code
      - Description
      - Price
    under each image.
    """
    pdf = FPDF(unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)

    images_per_page = 9
    cols = 3

    margin_x = 10
    cell_w = (210 - 2 * margin_x) / cols  # A4 width ~210mm
    img_h = 40
    text_h = 16
    top_y = 20

    for idx, row in df.iterrows():
        # New page every 9 items
        if idx % images_per_page == 0:
            pdf.add_page()
            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "Photo Catalogue", ln=1, align="C")
            pdf.ln(2)

        pos_in_page = idx % images_per_page
        row_idx = pos_in_page // cols
        col_idx = pos_in_page % cols

        x = margin_x + col_idx * cell_w
        y = top_y + row_idx * (img_h + text_h + 6)

        file = row["FILE"]
        code = str(row["CODE"]) if row["CODE"] else ""
        desc = str(row["DESCRIPTION"]) if row["DESCRIPTION"] else ""
        price = row["PRICE_INCL"]
        price_str = ""
        if price is not None and price != "":
            try:
                price_str = f"{float(price):.2f}"
            except Exception:
                price_str = str(price)

        # Add image
        try:
            img = Image.open(file).convert("RGB")
            img.thumbnail((int(cell_w - 10), img_h * 3))  # safe size

            tmp = tempfile.NamedTemporaryFile(
                dir=temp_dir, suffix=".jpg", delete=False
            )
            tmp_path = tmp.name
            tmp.close()  # close handle so FPDF can re-open it
            img.save(tmp_path, format="JPEG")

            pdf.image(tmp_path, x=x + 5, y=y, w=cell_w - 10)
        except Exception:
            pass

        # Text under image
        pdf.set_xy(x, y + img_h + 1)
        pdf.set_font("Arial", size=8)

        text_lines = []
        if code:
            text_lines.append(f"Code: {code}")
        if desc:
            text_lines.append(desc)
        if price_str:
            text_lines.append(f"Price: {price_str}")

        text_block = "\n".join(text_lines)
        pdf.multi_cell(cell_w, 4, text_block, 0, "L")

    # fpdf2 may return str, bytes, or bytearray depending on version
    res = pdf.output(dest="S")
    if isinstance(res, (bytes, bytearray)):
        pdf_bytes = bytes(res)
    else:
        pdf_bytes = res.encode("latin1")

    return pdf_bytes


# -----------------------------
# Streamlit UI
# -----------------------------

st.set_page_config(page_title="Photo Catalogue Builder", layout="wide")
st.title("📸 Photo Catalogue Builder")
st.write(
    "Upload your **photos** and a **price Excel file**. "
    "I'll match by code in the filename and build both an Excel and a PDF with "
    "Photo + Code + Description + Price."
)

st.markdown("### 1️⃣ Upload Photos")
uploaded_photos = st.file_uploader(
    "Select multiple product photos",
    type=["jpg", "jpeg", "png"],
    accept_multiple_files=True,
    help="Filenames must contain the product code, e.g. 8613900001-25pcs.jpg",
)

st.markdown("### 2️⃣ Upload Price Excel")
uploaded_price = st.file_uploader(
    "Upload the Excel price list",
    type=["xls", "xlsx"],
    help="I'll auto-detect Code, Description, and VAT-inclusive Price columns.",
)

if uploaded_photos and uploaded_price:
    if st.button("🔍 Match Photos & Build Catalogue"):
        with st.spinner("Processing... please wait"):
            try:
                # Read Excel
                price_df = pd.read_excel(uploaded_price)

                # Detect columns
                code_col = find_code_col(price_df)
                desc_col = find_desc_col(price_df)
                price_col = find_price_incl_col(price_df)

                st.caption(f"Using columns → CODE: `{code_col}`, DESCRIPTION: `{desc_col}`, PRICE INCL: `{price_col}`")

                # Build matched DF
                matched_df = build_matched_df(
                    price_df, uploaded_photos, code_col, desc_col, price_col
                )

                # Display summary table (no FILE col)
                display_df = matched_df[["FILENAME", "CODE", "DESCRIPTION", "PRICE_INCL"]].copy()
                display_df["PRICE_INCL"] = display_df["PRICE_INCL"].astype(str)
                st.markdown("### ✅ Matched Items Preview")
                st.dataframe(display_df, use_container_width=True)

                # Build Excel + PDF using temp dir
                with tempfile.TemporaryDirectory() as tmpdir:
                    excel_bytes = build_excel_with_thumbnails(matched_df, tmpdir)
                    pdf_bytes = build_pdf_grid(matched_df, tmpdir)

                st.markdown("### 📥 Downloads")

                st.download_button(
                    "⬇️ Download Excel with Thumbnails",
                    data=excel_bytes,
                    file_name="catalogue_with_photos.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                )

                st.download_button(
                    "⬇️ Download 3×3 PDF Catalogue",
                    data=pdf_bytes,
                    file_name="photo_catalogue.pdf",
                    mime="application/pdf",
                )

            except Exception as e:
                st.error(f"Something went wrong: {e}")

else:
    st.info("Please upload **photos** and a **price Excel file** to continue.")
